"""Load employee rows from LaoPayrollToolkit v5.xlsx (sheet EMPLOYEE_MASTER).

Prerequisites: ``lk_bank_master`` should include toolkit banks/branches (run
``seed_bank_master_from_toolkit`` first). ``lk_allowance_rates`` must contain
position-level allowance names used in the sheet.

Run inside API container:

  TOOLKIT_XLSX=/path/to/LaoPayrollToolkit\\\\ v5.xlsx \\
    docker compose exec api python -m app.db.seeds.seed_employee_master_from_toolkit

Default path: ``/toolkit/LaoPayrollToolkit v5.xlsx`` or repo ``documents/`` copy
(same resolution as ``seed_bank_master_from_toolkit``).

Upserts by ``employee_code``: existing rows are updated from the workbook;
new rows are inserted. ``created_by`` / ``updated_by`` are set to identify
the import.
"""
from __future__ import annotations

import os
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session, sessionmaker

from app.config import get_settings
from app.models.employee import Employee
from app.schemas.employee import EmployeeCreate

# EMPLOYEE_MASTER bank labels differ from LK_BankMaster / ``lk_bank_master`` (typo
# Agriculture→Agricultural, short BCEL/BIC names, ``-`` vs em dash in branch names).
# Values must match ``seed_bank_master_from_toolkit`` upserts exactly.
_EMP_BANK_BRANCH_TO_LK: dict[tuple[str, str], tuple[str, str]] = {
    ("Agriculture Promotion Bank (APB)", "APB Champasak Branch"): (
        "Agricultural Promotion Bank (APB)",
        "APB Champasak Branch",
    ),
    ("Agriculture Promotion Bank (APB)", "APB Head Office - Vientiane"): (
        "Agricultural Promotion Bank (APB)",
        "APB Head Office — Vientiane",
    ),
    ("Agriculture Promotion Bank (APB)", "APB Khammuane Branch"): (
        "Agricultural Promotion Bank (APB)",
        "APB Khammuane Branch",
    ),
    ("Agriculture Promotion Bank (APB)", "APB Luang Prabang Branch"): (
        "Agricultural Promotion Bank (APB)",
        "APB Luang Prabang Branch",
    ),
    ("Agriculture Promotion Bank (APB)", "APB Savannakhet Branch"): (
        "Agricultural Promotion Bank (APB)",
        "APB Savannakhet Branch",
    ),
    ("BCEL (Banque pour le Commerce Exterieur Lao)", "BCEL Bolikhamxay Branch"): (
        "Banque pour le Commerce Exterieur Lao Public (BCEL)",
        "BCEL Bolikhamxay / Pakxan Branch",
    ),
    ("BCEL (Banque pour le Commerce Exterieur Lao)", "BCEL Champasak Branch"): (
        "Banque pour le Commerce Exterieur Lao Public (BCEL)",
        "BCEL Champasak / Pakse Branch",
    ),
    ("BCEL (Banque pour le Commerce Exterieur Lao)", "BCEL Head Office - Vientiane"): (
        "Banque pour le Commerce Exterieur Lao Public (BCEL)",
        "BCEL Head Office — Vientiane",
    ),
    ("BCEL (Banque pour le Commerce Exterieur Lao)", "BCEL Khammuane Branch"): (
        "Banque pour le Commerce Exterieur Lao Public (BCEL)",
        "BCEL Khammuane / Thakhek Branch",
    ),
    ("BCEL (Banque pour le Commerce Exterieur Lao)", "BCEL Luang Prabang Branch"): (
        "Banque pour le Commerce Exterieur Lao Public (BCEL)",
        "BCEL Luang Prabang Branch",
    ),
    ("BCEL (Banque pour le Commerce Exterieur Lao)", "BCEL Oudomxay Branch"): (
        "Banque pour le Commerce Exterieur Lao Public (BCEL)",
        "BCEL Oudomxay / Muang Xay Branch",
    ),
    ("BCEL (Banque pour le Commerce Exterieur Lao)", "BCEL Savannakhet Branch"): (
        "Banque pour le Commerce Exterieur Lao Public (BCEL)",
        "BCEL Savannakhet Branch",
    ),
    ("BCEL (Banque pour le Commerce Exterieur Lao)", "BCEL Xiengkhuang Branch"): (
        "Banque pour le Commerce Exterieur Lao Public (BCEL)",
        "BCEL Xiengkhuang / Phonsavan Branch",
    ),
    ("BIC Bank Laos", "BIC Head Office - Vientiane"): (
        "BIC Bank Lao",
        "BIC Head Office — Vientiane",
    ),
    ("Joint Development Bank (JDB)", "JDB Vientiane Branch"): (
        "Joint Development Bank (JDB)",
        "JDB Head Office — Vientiane",
    ),
    ("Lao Development Bank (LDB)", "LDB Head Office - Vientiane"): (
        "Lao Development Bank (LDB)",
        "LDB Head Office — Vientiane",
    ),
    ("Lao Development Bank (LDB)", "LDB Pakse Branch"): (
        "Lao Development Bank (LDB)",
        "LDB Champasak / Pakse Branch",
    ),
}


def _resolve_unique_email(
    raw_email: str,
    employee_code: str,
    email_owner: dict[str, str],
) -> str:
    """Toolkit rows reuse the same gov.la address; the DB requires unique emails."""
    e = raw_email.lower().strip()
    if email_owner.get(e) == employee_code:
        return e
    if e not in email_owner:
        email_owner[e] = employee_code
        return e
    local, _, domain = e.partition("@")
    alt = f"{local}+{employee_code}@{domain}".lower()
    if alt not in email_owner:
        email_owner[alt] = employee_code
        return alt
    n = 2
    while True:
        alt2 = f"{local}+{employee_code}.{n}@{domain}".lower()
        if alt2 not in email_owner:
            email_owner[alt2] = employee_code
            return alt2
        n += 1


def _toolkit_path() -> Path:
    if env := os.environ.get("TOOLKIT_XLSX"):
        p = Path(env).expanduser().resolve()
        if p.is_file():
            return p
        raise SystemExit(f"TOOLKIT_XLSX not found: {p}")
    here = Path(__file__).resolve()
    candidates = [
        Path("/toolkit/LaoPayrollToolkit v5.xlsx"),
        Path("/tmp/LaoPayrollToolkit v5.xlsx"),
    ]
    if len(here.parents) > 5:
        candidates.append(here.parents[5] / "documents" / "LaoPayrollToolkit v5.xlsx")
    for p in candidates:
        if p.is_file():
            return p
    raise SystemExit(
        "LaoPayrollToolkit v5.xlsx not found. Set TOOLKIT_XLSX, or mount at /toolkit/LaoPayrollToolkit v5.xlsx"
    )


def _cell_date(val: object) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _str(v: object | None, max_len: int | None = None) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    if max_len:
        return s[:max_len]
    return s


def _int_cell(v: object) -> int:
    if v is None:
        return 0
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(round(v))
    if isinstance(v, Decimal):
        return int(v)
    s = str(v).strip()
    if not s:
        return 0
    return int(round(float(s)))


def _bool_cell(v: object) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("yes", "y", "true", "1"):
        return True
    return False


def _row_to_raw(row: tuple[object, ...], col: dict[str, int]) -> dict[str, object]:
    def g(key: str) -> object:
        i = col[key]
        return row[i] if i < len(row) else None

    dob = _cell_date(g("Date of Birth"))
    doj = _cell_date(g("Date of Joining"))
    if dob is None or doj is None:
        raise ValueError("Date of Birth and Date of Joining are required")

    fa = _str(g("Field Allowance Type"), 20)
    if fa is None or str(fa).strip().lower() == "none":
        field_allowance: str = "None"
    elif fa in ("Teaching", "Medical"):
        field_allowance = fa
    else:
        field_allowance = "None"

    sso = _str(g("Social Security No."), 12)
    swift = _str(g("SWIFT / BIC Code"), 12)

    bank_name = _str(g("Bank Name"), 70) or ""
    bank_branch = _str(g("Bank Branch"), 60) or ""
    if bank_name and bank_branch:
        mapped = _EMP_BANK_BRANCH_TO_LK.get((bank_name, bank_branch))
        if mapped:
            bank_name, bank_branch = mapped[0], mapped[1]

    return {
        "employee_code": _str(g("Employee Code"), 10),
        "title": _str(g("Title"), 10),
        "first_name": _str(g("First Name"), 80) or "",
        "last_name": _str(g("Last Name"), 80) or "",
        "gender": _str(g("Gender"), 6) or "Other",
        "date_of_birth": dob,
        "email": (_str(g("Email Address"), 100) or "").lower(),
        "mobile_number": _str(g("Mobile Number"), 20),
        "date_of_joining": doj,
        "employment_type": _str(g("Employment Type"), 15) or "Permanent",
        "position_title": _str(g("Position / Designation"), 100) or "",
        "education_level": _str(g("Education Qualification"), 40) or "",
        "prior_experience_years": max(0, min(40, _int_cell(g("Prior Experience (Yrs before Govt)")))),
        "grade": _int_cell(g("Grade")),
        "step": _int_cell(g("Step")),
        "civil_service_card_id": _str(g("Civil Service Card ID"), 12) or "",
        "sso_number": sso,
        "ministry_name": _str(g("Ministry / Province / Central Org"), 80) or "",
        "department_name": _str(g("Department"), 80) or "",
        "division_name": _str(g("Division"), 60),
        "service_country": _str(g("Service Country"), 30) or "Lao PDR",
        "service_province": _str(g("Service Province / Posting"), 60) or "",
        "service_district": _str(g("Service District"), 60),
        "profession_category": _str(g("Profession Category"), 20) or "",
        "is_remote_area": _bool_cell(g("Is Remote Area")),
        "is_foreign_posting": _bool_cell(g("Is Foreign Posting")),
        "is_hazardous_area": _bool_cell(g("Is Hazardous Area")),
        "house_no": _str(g("House No."), 30),
        "street": _str(g("Street"), 100),
        "area_baan": _str(g("Area / Baan"), 80),
        "province_of_residence": _str(g("Province of Residence"), 60),
        "pin_code": _str(g("PIN Code"), 10),
        "residence_country": _str(g("Country"), 60),
        "bank_name": bank_name,
        "bank_branch": bank_branch,
        "bank_branch_code": _str(g("Bank Branch Code"), 10),
        "bank_account_no": _str(g("Bank Account No."), 20) or "",
        "swift_code": swift,
        "has_spouse": _bool_cell(g("Has Spouse")),
        "eligible_children": max(0, min(3, _int_cell(g("No. of Eligible Children (max 3)")))),
        "position_level": _str(g("Position Level (for Allowance)"), 80) or "",
        "is_na_member": _bool_cell(g("Is NA Member")),
        "field_allowance_type": field_allowance,
        "is_active": True,
    }


def _header_map(headers: list[str]) -> dict[str, int]:
    col: dict[str, int] = {}
    for i, h in enumerate(headers):
        key = h.replace("\n", " ").strip()
        if key:
            col[key] = i
    return col


def main() -> None:
    path = _toolkit_path()
    wb = load_workbook(path, read_only=True, data_only=True)
    if "EMPLOYEE_MASTER" not in wb.sheetnames:
        raise SystemExit("Sheet EMPLOYEE_MASTER not found")
    ws = wb["EMPLOYEE_MASTER"]
    header_row = next(ws.iter_rows(min_row=3, max_row=3, values_only=True))
    headers = [str(c).replace("\n", " ").strip() if c is not None else "" for c in header_row]
    col = _header_map(headers)

    required_keys = (
        "Employee Code",
        "Title",
        "First Name",
        "Last Name",
        "Gender",
        "Date of Birth",
        "Email Address",
        "Date of Joining",
        "Employment Type",
        "Position / Designation",
        "Education Qualification",
        "Grade",
        "Step",
        "Civil Service Card ID",
        "Ministry / Province / Central Org",
        "Department",
        "Service Country",
        "Service Province / Posting",
        "Profession Category",
        "Bank Name",
        "Bank Branch",
        "Bank Account No.",
        "Position Level (for Allowance)",
    )
    for k in required_keys:
        if k not in col:
            raise SystemExit(f"Missing column: {k!r}")

    engine = create_engine(get_settings().alembic_database_url_sync)
    SessionLocal = sessionmaker(bind=engine, class_=Session)

    email_owner: dict[str, str] = {}
    with engine.connect() as conn:
        for code, em in conn.execute(text("SELECT employee_code, email FROM employee")):
            email_owner[str(em).lower()] = str(code)

    now = datetime.now(timezone.utc)
    label = "LaoPayrollToolkit EMPLOYEE_MASTER"
    imported = 0
    skipped = 0
    errors: list[tuple[int, str]] = []

    row_num = 3
    with SessionLocal() as session:
        for row in ws.iter_rows(min_row=4, values_only=True):
            row_num += 1
            existing: Employee | None = None
            cells = tuple(row)
            if not cells or cells[0] is None:
                continue
            if not str(cells[0]).strip():
                continue

            try:
                raw = _row_to_raw(cells, col)
                if not raw.get("employee_code"):
                    raise ValueError("Employee Code is empty")
                code = str(raw["employee_code"]).strip()
                existing = session.get(Employee, code)
                if existing:
                    old_em = existing.email.lower()
                    if email_owner.get(old_em) == code:
                        del email_owner[old_em]
                raw["email"] = _resolve_unique_email(str(raw["email"]), code, email_owner)
                body = EmployeeCreate.model_validate(raw)
            except (ValidationError, ValueError) as e:
                skipped += 1
                err = str(e)
                if isinstance(e, ValidationError):
                    err = e.errors().__repr__()
                errors.append((row_num, err[:500]))
                continue
            payload = body.model_dump()
            if existing:
                for k, v in payload.items():
                    setattr(existing, k, v)
                existing.updated_at = now
                existing.updated_by = label
            else:
                session.add(
                    Employee(
                        **payload,
                        created_by=label,
                        created_at=now,
                        uploaded_by_user_id=None,
                        owner_role="ROLE_ADMIN",
                    )
                )
            try:
                session.commit()
                imported += 1
            except Exception as ex:  # noqa: BLE001 — surface DB errors per row
                session.rollback()
                skipped += 1
                errors.append((row_num, str(ex)[:500]))

    wb.close()
    print(f"Done. Imported/updated: {imported}, skipped: {skipped}, path={path}")
    if errors:
        print("First errors (row, message):")
        for r, msg in errors[:25]:
            print(f"  row {r}: {msg}")
        if len(errors) > 25:
            print(f"  ... and {len(errors) - 25} more")


if __name__ == "__main__":
    main()
