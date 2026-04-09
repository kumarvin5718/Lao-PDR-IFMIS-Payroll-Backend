"""Employee list export (XLSX / PDF) — same scope and filters as list_employees + optional province."""

from __future__ import annotations

import os
import uuid
from datetime import date, datetime, timezone
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import String, and_, cast, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import Session
from weasyprint import HTML

from app.models.employee import Employee
from app.schemas.auth import User
from app.utils.scope import employee_scope_clause, employee_scope_clause_sync


def _fmt_date(d: date | None) -> str:
    if d is None:
        return ""
    return d.strftime("%d-%b-%Y")


def build_export_filter_lines(
    *,
    ministry: str | None,
    grade: int | None,
    province: str | None,
    employment_type: str | None,
    search: str | None,
    is_active: bool | None,
) -> list[str]:
    lines: list[str] = []
    if ministry:
        lines.append(f"Ministry: {ministry}")
    if grade is not None:
        lines.append(f"Grade: {grade}")
    if province:
        lines.append(f"Province: {province}")
    if employment_type:
        lines.append(f"Employment type: {employment_type}")
    if search:
        lines.append(f"Search: {search}")
    if is_active is True:
        lines.append("Status: Active")
    elif is_active is False:
        lines.append("Status: Inactive")
    else:
        lines.append("Status: All")
    return lines


async def build_export_where_async(
    db: AsyncSession,
    current_user: User,
    *,
    ministry: str | None,
    grade: int | None,
    province: str | None,
    employment_type: str | None,
    search: str | None,
    is_active: bool | None,
) -> Any:
    scope = await employee_scope_clause(db, current_user)
    parts: list = []
    if scope is not None:
        parts.append(scope)

    if current_user.role == "ROLE_ADMIN" and ministry and ministry.strip():
        parts.append(Employee.ministry_name == ministry.strip())

    if grade is not None:
        parts.append(Employee.grade == grade)
    if employment_type is not None and employment_type.strip():
        parts.append(Employee.employment_type == employment_type.strip())
    if province is not None and province.strip():
        parts.append(Employee.service_province == province.strip())
    if is_active is not None:
        parts.append(Employee.is_active == is_active)

    if search is not None and len(search.strip()) >= 2:
        term = f"%{search.strip()}%"
        grade_str = cast(Employee.grade, String)
        parts.append(
            or_(
                Employee.employee_code.ilike(term),
                Employee.first_name.ilike(term),
                Employee.last_name.ilike(term),
                Employee.civil_service_card_id.ilike(term),
                grade_str.ilike(term),
            ),
        )

    if not parts:
        return None
    return and_(*parts)


def build_export_where_sync(
    db: Session,
    current_user: User,
    *,
    ministry: str | None,
    grade: int | None,
    province: str | None,
    employment_type: str | None,
    search: str | None,
    is_active: bool | None,
) -> Any:
    scope = employee_scope_clause_sync(db, current_user)
    parts: list = []
    if scope is not None:
        parts.append(scope)

    if current_user.role == "ROLE_ADMIN" and ministry and ministry.strip():
        parts.append(Employee.ministry_name == ministry.strip())

    if grade is not None:
        parts.append(Employee.grade == grade)
    if employment_type is not None and employment_type.strip():
        parts.append(Employee.employment_type == employment_type.strip())
    if province is not None and province.strip():
        parts.append(Employee.service_province == province.strip())
    if is_active is not None:
        parts.append(Employee.is_active == is_active)

    if search is not None and len(search.strip()) >= 2:
        term = f"%{search.strip()}%"
        grade_str = cast(Employee.grade, String)
        parts.append(
            or_(
                Employee.employee_code.ilike(term),
                Employee.first_name.ilike(term),
                Employee.last_name.ilike(term),
                Employee.civil_service_card_id.ilike(term),
                grade_str.ilike(term),
            ),
        )

    if not parts:
        return None
    return and_(*parts)


async def count_export_rows(
    db: AsyncSession,
    current_user: User,
    *,
    ministry: str | None,
    grade: int | None,
    province: str | None,
    employment_type: str | None,
    search: str | None,
    is_active: bool | None,
) -> int:
    wc = await build_export_where_async(
        db,
        current_user,
        ministry=ministry,
        grade=grade,
        province=province,
        employment_type=employment_type,
        search=search,
        is_active=is_active,
    )
    stmt = select(func.count()).select_from(Employee)
    if wc is not None:
        stmt = stmt.where(wc)
    return int(await db.scalar(stmt) or 0)


async def fetch_export_rows(
    db: AsyncSession,
    current_user: User,
    *,
    ministry: str | None,
    grade: int | None,
    province: str | None,
    employment_type: str | None,
    search: str | None,
    is_active: bool | None,
    limit: int | None = None,
) -> list[Employee]:
    wc = await build_export_where_async(
        db,
        current_user,
        ministry=ministry,
        grade=grade,
        province=province,
        employment_type=employment_type,
        search=search,
        is_active=is_active,
    )
    stmt = select(Employee).order_by(Employee.employee_code.asc())
    if wc is not None:
        stmt = stmt.where(wc)
    if limit is not None:
        stmt = stmt.limit(limit)
    result = await db.execute(stmt)
    return list(result.scalars().all())


XLSX_HEADERS = [
    "Employee Code",
    "Title",
    "First Name",
    "Last Name",
    "Gender",
    "Date of Birth",
    "Email",
    "Mobile",
    "Date of Joining",
    "Employment Type",
    "Position",
    "Education",
    "Prior Exp Years",
    "Grade",
    "Step",
    "Civil Service Card ID",
    "SSO Number",
    "Ministry",
    "Department",
    "Division",
    "Service Country",
    "Service Province",
    "Service District",
    "Profession Category",
    "Is Remote",
    "Is Foreign",
    "Is Hazardous",
    "House No",
    "Street",
    "Area/Baan",
    "Province of Residence",
    "PIN Code",
    "Bank Name",
    "Bank Branch",
    "Bank Account No",
    "Has Spouse",
    "Eligible Children",
    "Position Level",
    "Is NA Member",
    "Field Allowance Type",
    "Is Active",
]


def _row_values(e: Employee) -> list[Any]:
    return [
        e.employee_code,
        e.title,
        e.first_name,
        e.last_name,
        e.gender,
        _fmt_date(e.date_of_birth),
        e.email,
        e.mobile_number or "",
        _fmt_date(e.date_of_joining),
        e.employment_type,
        e.position_title,
        e.education_level,
        e.prior_experience_years,
        e.grade,
        e.step,
        e.civil_service_card_id,
        e.sso_number or "",
        e.ministry_name,
        e.department_name,
        e.division_name or "",
        e.service_country,
        e.service_province,
        e.service_district or "",
        e.profession_category,
        e.is_remote_area,
        e.is_foreign_posting,
        e.is_hazardous_area,
        e.house_no or "",
        e.street or "",
        e.area_baan or "",
        e.province_of_residence or "",
        e.pin_code or "",
        e.bank_name,
        e.bank_branch,
        e.bank_account_no,
        e.has_spouse,
        e.eligible_children,
        e.position_level,
        e.is_na_member,
        e.field_allowance_type,
        e.is_active,
    ]


def write_employees_xlsx(rows: list[Employee], path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Employees"
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="BDD7EE")

    ws.append(XLSX_HEADERS)
    for c in range(1, len(XLSX_HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for e in rows:
        ws.append(_row_values(e))

    ws.freeze_panes = "A2"

    for col_idx in range(1, len(XLSX_HEADERS) + 1):
        max_len = 10
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.value is None:
                    continue
                max_len = min(40, max(max_len, len(str(cell.value))))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len

    wb.save(path)


def write_employees_pdf(
    rows: list[Employee],
    path: str,
    *,
    filter_lines: list[str],
    generated_at: datetime,
) -> None:
    gen = generated_at.strftime("%d-%b-%Y %H:%M UTC")
    filter_html = "<br/>".join(filter_lines) if filter_lines else "No filters"
    table_rows = []
    for e in rows:
        full_name = f"{e.title} {e.first_name} {e.last_name}".strip()
        table_rows.append(
            "<tr>"
            f"<td>{e.employee_code}</td>"
            f"<td>{full_name}</td>"
            f"<td>{e.ministry_name}</td>"
            f"<td>{e.department_name}</td>"
            f"<td>G{e.grade}/S{e.step}</td>"
            f"<td>{e.employment_type}</td>"
            f"<td>{e.service_province}</td>"
            f"<td>{'Yes' if e.is_active else 'No'}</td>"
            "</tr>"
        )
    body = "\n".join(table_rows)
    html = f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8"/>
<style>
@page {{
  size: A4 landscape;
  margin: 12mm;
  @bottom-right {{
    content: "Page " counter(page) " of " counter(pages);
    font-size: 8pt;
  }}
}}
body {{ font-family: DejaVu Sans, sans-serif; font-size: 9pt; }}
h1 {{ font-size: 14pt; margin: 0 0 4px 0; }}
.sub {{ font-size: 9pt; color: #333; margin-bottom: 8px; }}
table {{ border-collapse: collapse; width: 100%; }}
th, td {{ border: 1px solid #ccc; padding: 3px 5px; text-align: left; }}
th {{ background: #BDD7EE; font-weight: bold; font-size: 8pt; }}
</style>
</head>
<body>
<h1>IFMS Payroll — Employee List</h1>
<div class="sub">Filters: {filter_html}<br/>Generated: {gen}</div>
<table>
<thead>
<tr>
<th>Employee Code</th><th>Full Name</th><th>Ministry</th><th>Department</th>
<th>Grade/Step</th><th>Employment Type</th><th>Province</th><th>Is Active</th>
</tr>
</thead>
<tbody>
{body}
</tbody>
</table>
</body>
</html>
"""
    HTML(string=html).write_pdf(path)


def export_employees_xlsx_sync(
    db: Session,
    current_user: User,
    *,
    ministry: str | None,
    grade: int | None,
    province: str | None,
    employment_type: str | None,
    search: str | None,
    is_active: bool | None,
    reports_dir: str,
) -> dict[str, Any]:
    wc = build_export_where_sync(
        db,
        current_user,
        ministry=ministry,
        grade=grade,
        province=province,
        employment_type=employment_type,
        search=search,
        is_active=is_active,
    )
    stmt = select(Employee).order_by(Employee.employee_code.asc())
    if wc is not None:
        stmt = stmt.where(wc)
    rows = list(db.execute(stmt).scalars().all())
    os.makedirs(reports_dir, exist_ok=True)
    filename = f"employees_export_{uuid.uuid4().hex[:12]}.xlsx"
    path = os.path.join(reports_dir, filename)
    write_employees_xlsx(rows, path)
    return {"file_path": path, "rows": len(rows)}
