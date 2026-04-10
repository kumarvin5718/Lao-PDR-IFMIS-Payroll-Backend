"""Load LK_BankMaster from LaoPayrollToolkit v5.xlsx (sheet LK_BankMaster).

Run inside API container (openpyxl required), with toolkit path mounted or copied:

  TOOLKIT_XLSX=/path/to/LaoPayrollToolkit\\ v5.xlsx \\
    docker compose exec api python -m app.db.seeds.seed_bank_master_from_toolkit

Default path: <IFMS>/documents/LaoPayrollToolkit v5.xlsx (five levels above this file).
"""
from __future__ import annotations

import os
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import create_engine, text

from app.config import get_settings


def _toolkit_path() -> Path:
    if env := os.environ.get("TOOLKIT_XLSX"):
        p = Path(env).expanduser().resolve()
        if p.is_file():
            return p
        raise SystemExit(f"TOOLKIT_XLSX not found: {p}")
    here = Path(__file__).resolve()
    candidates = [
        Path("/toolkit/LaoPayrollToolkit v5.xlsx"),
        Path("/tmp/LaoPayrollToolkit v5.xlsx"),
    ]
    if len(here.parents) > 5:
        candidates.append(here.parents[5] / "documents" / "LaoPayrollToolkit v5.xlsx")
    for p in candidates:
        if p.is_file():
            return p
    raise SystemExit(
        "LaoPayrollToolkit v5.xlsx not found. Set TOOLKIT_XLSX, or mount the file at /toolkit/LaoPayrollToolkit v5.xlsx"
    )


def _cell_date(val: object) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _norm_swift(v: object) -> str:
    s = (str(v).strip() if v is not None else "") or "000000000000"
    return (s + "XXXXXXXXXXXX")[:12]


def _str(v: object | None, max_len: int | None = None) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    if max_len:
        return s[:max_len]
    return s


def main() -> None:
    path = _toolkit_path()
    if not path.is_file():
        raise SystemExit(f"Toolkit not found: {path}. Set TOOLKIT_XLSX or copy LaoPayrollToolkit v5.xlsx into documents/.")

    wb = load_workbook(path, read_only=True, data_only=True)
    if "LK_BankMaster" not in wb.sheetnames:
        raise SystemExit("Sheet LK_BankMaster not found")
    ws = wb["LK_BankMaster"]

    # Row 4 headers: A=# B=Bank Name C=Abbrev D=Bank Key E=Category F=Branch Name G=Branch Code H=City I=SWIFT ...
    rows_out: list[dict] = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        cells = list(row)
        if len(cells) < 9:
            continue
        idx = cells[0]
        if idx is None:
            continue
        if not str(idx).strip().isdigit():
            continue
        bank_name = _str(cells[1], 70)
        if not bank_name:
            continue
        bank_key_raw = _str(cells[3], 6) or "XXXXXX"
        bank_key = bank_key_raw[:6].ljust(6, "0")[:6]
        branch_name = _str(cells[5], 60) or "Main"
        branch_code = _str(cells[6], 10) or "UNKN"
        swift = _norm_swift(cells[8] if len(cells) > 8 else None)

        eff_from = _cell_date(cells[15]) if len(cells) > 15 else None
        eff_to = _cell_date(cells[16]) if len(cells) > 16 else None
        last_upd = _cell_date(cells[17]) if len(cells) > 17 else None

        rows_out.append(
            {
                "bank_name": bank_name,
                "bank_key": bank_key,
                "branch_name": branch_name,
                "branch_code": branch_code,
                "swift_code": swift,
                "category": _str(cells[4], 120),
                "bank_abbrev": _str(cells[2], 20),
                "city": _str(cells[7], 100),
                "branch_address": _str(cells[9], 500) if len(cells) > 9 else None,
                "bank_hq_address": _str(cells[10], 500) if len(cells) > 10 else None,
                "telephone": _str(cells[11], 80) if len(cells) > 11 else None,
                "ownership": _str(cells[12], 300) if len(cells) > 12 else None,
                "established": _str(cells[13], 30) if len(cells) > 13 else None,
                "website": _str(cells[14], 300) if len(cells) > 14 else None,
                "effective_from": eff_from,
                "effective_to": eff_to,
                "last_updated": last_upd,
                "last_updated_by": _str(cells[18], 80) if len(cells) > 18 else None,
                "circular_ref": _str(cells[19], 80) if len(cells) > 19 else None,
                "change_remarks": _str(cells[20], 200) if len(cells) > 20 else None,
            }
        )

    wb.close()
    if not rows_out:
        raise SystemExit("No data rows parsed from LK_BankMaster")

    engine = create_engine(get_settings().alembic_database_url_sync)
    sql = text(
        """
        INSERT INTO lk_bank_master (
            bank_name, bank_key, branch_name, branch_code, swift_code,
            category, bank_abbrev, city, branch_address, bank_hq_address,
            telephone, ownership, established, website,
            effective_from, effective_to, last_updated, last_updated_by,
            circular_ref, change_remarks
        ) VALUES (
            :bank_name, :bank_key, :branch_name, :branch_code, :swift_code,
            :category, :bank_abbrev, :city, :branch_address, :bank_hq_address,
            :telephone, :ownership, :established, :website,
            :effective_from, :effective_to, :last_updated, :last_updated_by,
            :circular_ref, :change_remarks
        )
        ON CONFLICT (bank_name, branch_name) DO UPDATE SET
            bank_key = EXCLUDED.bank_key,
            branch_code = EXCLUDED.branch_code,
            swift_code = EXCLUDED.swift_code,
            category = EXCLUDED.category,
            bank_abbrev = EXCLUDED.bank_abbrev,
            city = EXCLUDED.city,
            branch_address = EXCLUDED.branch_address,
            bank_hq_address = EXCLUDED.bank_hq_address,
            telephone = EXCLUDED.telephone,
            ownership = EXCLUDED.ownership,
            established = EXCLUDED.established,
            website = EXCLUDED.website,
            effective_from = EXCLUDED.effective_from,
            effective_to = EXCLUDED.effective_to,
            last_updated = EXCLUDED.last_updated,
            last_updated_by = EXCLUDED.last_updated_by,
            circular_ref = EXCLUDED.circular_ref,
            change_remarks = EXCLUDED.change_remarks
        """
    )

    with engine.begin() as conn:
        for r in rows_out:
            conn.execute(sql, r)

    print(f"Upserted {len(rows_out)} bank master rows from {path}")


if __name__ == "__main__":
    main()
