import io
import json
import uuid
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from uuid import UUID

import openpyxl
from fastapi import HTTPException, UploadFile, status
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import ValidationError
from sqlalchemy import func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import get_settings
from app.models.employee import Employee
from app.models.upload_session import UploadSession, UploadSessionRow
from app.schemas.auth import User
from app.schemas.employee import EmployeeCreate

EXPECTED_COLUMNS: list[str] = [
    "employee_code",
    "title",
    "first_name",
    "last_name",
    "gender",
    "date_of_birth",
    "email",
    "mobile_number",
    "date_of_joining",
    "employment_type",
    "position_title",
    "education_level",
    "prior_experience_years",
    "grade",
    "step",
    "civil_service_card_id",
    "sso_number",
    "ministry_name",
    "department_name",
    "division_name",
    "service_country",
    "service_province",
    "service_district",
    "profession_category",
    "is_remote_area",
    "is_foreign_posting",
    "is_hazardous_area",
    "house_no",
    "street",
    "area_baan",
    "province_of_residence",
    "pin_code",
    "residence_country",
    "bank_name",
    "bank_branch",
    "bank_branch_code",
    "bank_account_no",
    "swift_code",
    "has_spouse",
    "eligible_children",
    "position_level",
    "is_na_member",
    "field_allowance_type",
    "is_active",
]

_BOOL_COLS = frozenset(
    {
        "is_remote_area",
        "is_foreign_posting",
        "is_hazardous_area",
        "has_spouse",
        "is_na_member",
        "is_active",
    }
)
_DATE_COLS = frozenset({"date_of_birth", "date_of_joining"})
_INT_COLS = frozenset({"prior_experience_years", "grade", "step", "eligible_children"})


def _get_upload_dir() -> Path:
    settings = get_settings()
    d = Path(getattr(settings, "UPLOAD_DIR", "/app/uploads"))
    d.mkdir(parents=True, exist_ok=True)
    return d


def _parse_bool(val: object) -> bool:
    if isinstance(val, bool):
        return val
    if isinstance(val, int):
        return bool(val)
    if isinstance(val, str):
        return val.strip().lower() in ("true", "yes", "1", "y")
    return False


def _parse_date(val: object) -> str | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel

            dt = from_excel(float(val))
            if isinstance(dt, datetime):
                return dt.date().isoformat()
            if isinstance(dt, date):
                return dt.isoformat()
        except Exception:
            return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                continue
        return None
    return None


def _to_int(val: object) -> int | None:
    if val is None or val == "":
        return None
    if isinstance(val, bool):
        return None
    try:
        if isinstance(val, float):
            return int(val)
        return int(str(val).strip())
    except (ValueError, TypeError):
        return None


def _format_validation_errors(exc: ValidationError) -> list[str]:
    out: list[str] = []
    for err in exc.errors():
        loc = ".".join(str(x) for x in err.get("loc", ()))
        msg = err.get("msg", "invalid")
        out.append(f"{loc}: {msg}" if loc else str(msg))
    return out


def _row_to_dict(row_values: list[object], col_names: list[str]) -> dict[str, object]:
    row: dict[str, object] = {}
    for i, name in enumerate(col_names):
        cell = row_values[i] if i < len(row_values) else None
        if name in _BOOL_COLS:
            row[name] = _parse_bool(cell)
        elif name in _DATE_COLS:
            row[name] = _parse_date(cell)
        elif name in _INT_COLS:
            row[name] = _to_int(cell)
        elif cell is None or cell == "":
            row[name] = None
        elif isinstance(cell, str):
            row[name] = cell.strip() if cell.strip() else None
        else:
            row[name] = cell

    ft = row.get("field_allowance_type")
    if ft is None or (isinstance(ft, str) and ft.strip() == ""):
        row["field_allowance_type"] = "None"
    elif isinstance(ft, str) and ft.strip().lower() == "none":
        row["field_allowance_type"] = "None"

    return row


async def parse_upload(
    db: AsyncSession,
    file: UploadFile,
    current_user: User,
) -> dict:
    filename = (file.filename or "").lower()
    if filename.endswith(".xlsm"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"code": "ERR_UPLOAD_MACRO_DETECTED", "message": "Macro-enabled workbooks are not allowed"},
        )
    if not filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"code": "ERR_UPLOAD_INVALID_FORMAT", "message": "Only .xlsx files are accepted"},
        )

    content = await file.read()
    max_bytes = 10 * 1024 * 1024
    if len(content) > max_bytes:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail={"code": "ERR_UPLOAD_TOO_LARGE", "message": "File exceeds 10 MB"},
        )

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"code": "ERR_UPLOAD_INVALID_FORMAT", "message": f"Could not read workbook: {exc!s}"},
        ) from exc

    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    if headers != EXPECTED_COLUMNS:
        missing = sorted(set(EXPECTED_COLUMNS) - set(headers))
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "code": "ERR_UPLOAD_INVALID_FORMAT",
                "message": f"Missing columns: {missing}" if missing else "Header row must match the template exactly (column order and names)",
            },
        )

    file_name = f"{uuid.uuid4()}.xlsx"
    path = _get_upload_dir() / file_name
    path.write_bytes(content)

    session_id = uuid.uuid4()
    session = UploadSession(
        session_id=session_id,
        upload_type="EMPLOYEE",
        uploaded_by=current_user.full_name,
        file_path=str(path),
        status="PARSING",
        expires_at=datetime.now(timezone.utc) + timedelta(minutes=30),
    )
    db.add(session)
    await db.commit()
    await db.refresh(session)

    all_row_objs: list[UploadSessionRow] = []
    ok_rows = 0
    error_rows = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None:
            continue
        row_list = list(row)
        while len(row_list) < len(EXPECTED_COLUMNS):
            row_list.append(None)
        row_list = row_list[: len(EXPECTED_COLUMNS)]
        if all(v is None or v == "" for v in row_list):
            continue

        row_dict = _row_to_dict(row_list, EXPECTED_COLUMNS)
        errors: list[str] = []
        row_status = "OK"

        try:
            EmployeeCreate.model_validate(row_dict)
        except ValidationError as exc:
            errors = _format_validation_errors(exc)
            row_status = "ERROR"

        if row_status == "OK":
            ok_rows += 1
        else:
            error_rows += 1

        ec = row_dict.get("employee_code")
        emp_code: str | None
        if ec is None or ec == "":
            emp_code = None
        else:
            emp_code = str(ec).strip() or None

        raw_json = json.dumps(row_dict, default=str)
        row_obj = UploadSessionRow(
            session_id=session.session_id,
            row_number=row_idx,
            employee_code=emp_code,
            raw_data=raw_json,
            status=row_status,
            errors=json.dumps(errors) if errors else None,
            warnings=None,
        )
        db.add(row_obj)
        all_row_objs.append(row_obj)

    total = len(all_row_objs)
    final_status = "PARSED" if error_rows == 0 else "PARSED_WITH_ERRORS"
    session.status = final_status
    session.total_rows = total
    session.valid_rows = ok_rows
    session.warning_rows = 0
    session.error_rows = error_rows

    await db.commit()

    def row_to_preview(r: UploadSessionRow) -> dict:
        return {
            "row_number": r.row_number,
            "employee_code": r.employee_code,
            "status": r.status,
            "errors": json.loads(r.errors) if r.errors else [],
            "warnings": json.loads(r.warnings) if r.warnings else [],
        }

    return {
        "session_id": str(session.session_id),
        "total_rows": total,
        "valid_rows": ok_rows,
        "error_rows": error_rows,
        "warning_rows": 0,
        "can_commit": error_rows == 0,
        "preview": [row_to_preview(r) for r in all_row_objs[:5]],
        "all_rows": [row_to_preview(r) for r in all_row_objs],
    }


def _aware_expiry(exp: datetime) -> datetime:
    if exp.tzinfo is None:
        return exp.replace(tzinfo=timezone.utc)
    return exp


async def commit_upload(
    db: AsyncSession,
    session_id: UUID,
    current_user: User,
) -> dict:
    res = await db.execute(select(UploadSession).where(UploadSession.session_id == session_id))
    session = res.scalar_one_or_none()
    if session is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"code": "ERR_UPLOAD_SESSION_NOT_FOUND", "message": "Upload session not found"},
        )

    now = datetime.now(timezone.utc)
    if _aware_expiry(session.expires_at) < now:
        raise HTTPException(
            status_code=status.HTTP_410_GONE,
            detail={"code": "ERR_UPLOAD_SESSION_EXPIRED", "message": "Upload session has expired"},
        )

    if session.status not in ("PARSED", "PARSED_WITH_ERRORS"):
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail={
                "code": "ERR_UPLOAD_SESSION_NOT_FOUND",
                "message": "Session not in parseable state",
            },
        )

    err_count_res = await db.execute(
        select(func.count())
        .select_from(UploadSessionRow)
        .where(
            UploadSessionRow.session_id == session_id,
            UploadSessionRow.status == "ERROR",
        )
    )
    error_row_count = int(err_count_res.scalar_one() or 0)

    rows_res = await db.execute(
        select(UploadSessionRow)
        .where(
            UploadSessionRow.session_id == session_id,
            UploadSessionRow.status != "ERROR",
        )
        .order_by(UploadSessionRow.row_number)
    )
    rows = rows_res.scalars().all()

    committed = 0
    skipped_duplicate = 0

    for row in rows:
        raw = json.loads(row.raw_data)
        try:
            emp_data = EmployeeCreate.model_validate(raw)
        except ValidationError:
            continue

        dup_res = await db.execute(
            select(func.count()).select_from(Employee).where(Employee.employee_code == emp_data.employee_code)
        )
        if int(dup_res.scalar_one() or 0) > 0:
            skipped_duplicate += 1
            continue

        emp = Employee(
            **emp_data.model_dump(),
            created_by=current_user.full_name,
            created_at=datetime.now(timezone.utc),
        )
        try:
            async with db.begin_nested():
                db.add(emp)
                await db.flush()
                committed += 1
        except IntegrityError:
            skipped_duplicate += 1

    session.status = "COMMITTED"
    await db.commit()

    return {
        "session_id": str(session_id),
        "committed": committed,
        "skipped_duplicates": skipped_duplicate,
        "skipped_errors": error_row_count,
    }


def get_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Upload Template"

    header_fill = PatternFill(fill_type="solid", fgColor="1B3A6B")
    header_font = Font(bold=True, color="FFFFFFFF")

    for col_idx, name in enumerate(EXPECTED_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font

    example: list[object] = [
        "LAO00001",
        "Mr.",
        "Somchai",
        "Vongsa",
        "Male",
        "1985-06-15",
        "somchai.vongsa@mof.gov.la",
        "+85620xxxxxxx",
        "2010-03-01",
        "Permanent",
        "Senior Analyst",
        "Bachelor's Degree",
        2,
        4,
        3,
        "123456789012",
        "123456789012",
        "Ministry of Finance",
        "Budget Department",
        "Planning",
        "Lao PDR",
        "Vientiane Capital",
        "Chanthabouly",
        "General",
        False,
        False,
        False,
        "123",
        "Lane Xang Ave",
        "Baan Nongbone",
        "Vientiane Capital",
        "01000",
        "Lao PDR",
        "BCEL",
        "Main Branch",
        "001",
        "0301010XXXXXXX",
        "BCELLALAVX",
        True,
        2,
        "Grade 4",
        False,
        "None",
        True,
    ]

    for col_idx, val in enumerate(example, start=1):
        ws.cell(row=2, column=col_idx, value=val)

    for col_idx, name in enumerate(EXPECTED_COLUMNS, start=1):
        ex_val = example[col_idx - 1]
        w = max(len(name), len(str(ex_val))) + 4
        ws.column_dimensions[get_column_letter(col_idx)].width = min(w, 60)

    last_data_row = 1000
    validations: list[tuple[str, str]] = [
        ("title", '"Mr.,Ms.,Mrs.,Dr.,Prof."'),
        ("gender", '"Male,Female"'),
        ("employment_type", '"Permanent,Probationary,Contract,Intern"'),
        ("is_remote_area", '"True,False"'),
        ("is_foreign_posting", '"True,False"'),
        ("is_hazardous_area", '"True,False"'),
        ("has_spouse", '"True,False"'),
        ("is_na_member", '"True,False"'),
        ("field_allowance_type", '"Teaching,Medical,None"'),
        ("is_active", '"True,False"'),
    ]
    for col_name, formula in validations:
        col_idx = EXPECTED_COLUMNS.index(col_name) + 1
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}3:{col_letter}{last_data_row}")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
